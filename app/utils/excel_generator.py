"""
Generador de Excel
Crea reportes en Excel usando openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
import os
from typing import List, Dict
from ..schemas import FiltrosReporteProyectos


class ExcelGenerator:
    """Generador de reportes Excel"""
    
    def __init__(self):
        # Crear carpeta para reportes si no existe
        self.output_dir = "reportes_excel"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generar_reporte_proyectos(
        self,
        datos: List[Dict],
        filtros: FiltrosReporteProyectos
    ) -> str:
        """
        Generar Excel de reporte de proyectos
        
        Returns:
            Path del Excel generado
        """
        # Nombre del archivo
        filename = f"reporte_proyectos_{date.today()}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyectos"
        
        # Estilos
        header_fill = PatternFill(start_color="A10000", end_color="A10000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "📊 REPORTE DE PROYECTOS - PORTAFOLIO DEVS"
        ws['A1'].font = Font(bold=True, size=16, color="A10000")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información
        ws['A3'] = f"Fecha de generación: {date.today()}"
        if filtros.id_programador:
            ws['A4'] = f"Programador: {filtros.id_programador}"
        if filtros.tipo:
            ws['A5'] = f"Tipo: {filtros.tipo}"
        
        # Estadísticas
        row = 7
        ws[f'A{row}'] = "📈 ESTADÍSTICAS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        total = len(datos)
        academicos = len([d for d in datos if d.get('tipo') == 'academico'])
        laborales = len([d for d in datos if d.get('tipo') == 'laboral'])
        
        ws[f'A{row}'] = "Total de proyectos:"
        ws[f'B{row}'] = total
        row += 1
        ws[f'A{row}'] = "Proyectos académicos:"
        ws[f'B{row}'] = academicos
        row += 1
        ws[f'A{row}'] = "Proyectos laborales:"
        ws[f'B{row}'] = laborales
        
        # Tabla de proyectos
        row += 3
        ws[f'A{row}'] = "📋 DETALLE DE PROYECTOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 2
        # Encabezados
        headers = ['ID', 'Nombre', 'Tipo', 'Participación', 'Tecnologías', 'Repo URL']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        for proyecto in datos:
            row += 1
            # ID
            ws.cell(row=row, column=1, value=proyecto.get('id', 'N/A'))
            # Nombre
            ws.cell(row=row, column=2, value=proyecto.get('nombre', 'N/A'))
            # Tipo
            ws.cell(row=row, column=3, value=proyecto.get('tipo', 'N/A'))
            # Participación
            ws.cell(row=row, column=4, value=proyecto.get('participacion', 'N/A'))
            # Tecnologías
            techs = proyecto.get('tecnologias', [])
            if isinstance(techs, list):
                ws.cell(row=row, column=5, value=', '.join(techs))
            else:
                ws.cell(row=row, column=5, value=str(techs))
            # Repo URL
            ws.cell(row=row, column=6, value=proyecto.get('repoUrl', 'N/A'))
            
            # Aplicar bordes
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        
        # Guardar
        wb.save(filepath)
        print(f"✅ Excel generado: {filepath}")
        
        return filepath